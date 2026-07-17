from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.auth import get_current_user, financials_visible
from app.models.inventory import (
    InventoryItem, InventoryAdjustment,
    InventoryCategory, AdjustmentReason,
)
from app.models.settings import AppSetting

router = APIRouter(prefix="/inventory", tags=["inventory"])
templates = Jinja2Templates(directory="app/templates")

CATEGORY_LABELS = {
    # Carbon Steel — Tube & Pipe
    "steel_pipe":           "Steel — Pipe (Sched 40)",
    "steel_rect_tube":      "Steel — Rect Tube",
    "steel_sq_tube":        "Steel — Sq Tube",
    "steel_dom_tube":       "Steel — DOM Tube",
    # Carbon Steel — Structural
    "steel_channel":        "Steel — Channel",
    "steel_angle":          "Steel — Angle (Equal)",
    "steel_angle_unequal":  "Steel — Angle (Unequal)",
    "steel_ibeam":          "Steel — I-Beam",
    "steel_wide_flange":    "Steel — Wide Flange",
    "steel_tstock":         "Steel — T-Stock",
    "steel_columns":        "Steel — Columns",
    # Carbon Steel — Bar
    "steel_flat_bar":       "Steel — Flat Bar",
    "steel_round_bar":      "Steel — Round Bar",
    "steel_square_bar":     "Steel — Square Bar",
    "steel_strip_hr":       "Steel — HR Strip",
    # Carbon Steel — Plate & Sheet
    "steel_plate_a36":      "Steel — Plate A36",
    "steel_plate_ar400":    "Steel — Plate AR400",
    "steel_floor_plate":    "Steel — Floor Plate",
    "steel_sheet_hr":       "Steel — Sheet HR",
    "steel_sheet_cr":       "Steel — Sheet CR",
    "steel_sheet_galv":     "Steel — Sheet Galvanized",
    "steel_sheet_perf":     "Steel — Sheet Perforated",
    # Carbon Steel — Misc Shapes
    "steel_expanded":       "Steel — Expanded Metal",
    "steel_grip_strut":     "Steel — Grip Strut",
    "steel_bar_grating":    "Steel — Bar Grating",
    "steel_decking":        "Steel — Floor Decking",
    "steel_rebar":          "Steel — Rebar",
    # Aluminum — Structural
    "alum_angle":           "Alum — Angle (Equal)",
    "alum_angle_unequal":   "Alum — Angle (Unequal)",
    "alum_channel":         "Alum — Channel",
    "alum_flat_bar":        "Alum — Flat Bar",
    "alum_round":           "Alum — Round Bar",
    "alum_square_bar":      "Alum — Square Bar",
    "alum_sq_tube":         "Alum — Sq Tube",
    "alum_grip_strut":      "Alum — Grip Strut",
    # Aluminum — Sheet
    "alum_sheet":           "Alum — Sheet",
    "alum_treadbrite":      "Alum — Tread Brite",
    # Stainless
    "ss_round_bar":         "Stainless — Round Bar",
    "ss_square_bar":        "Stainless — Square Bar",
    "ss_sheet":             "Stainless — Sheet",
    # Hardware
    "hardware_fasteners":   "Hardware — Fasteners",
    "hardware_caps":        "Hardware — Caps",
    "hardware_gussets":     "Hardware — Gussets",
    "hardware_base_plates": "Hardware — Base Plates",
    "hardware_handrail":    "Hardware — Handrail",
    "hardware_hinges":      "Hardware — Hinges",
    # Other
    "bumper_posts":         "Bumper Posts",
    "consumables":          "Consumables",
    "retail":               "Retail / Walk-In",
}

REASON_LABELS = {
    "received":   "Received from Supplier",
    "used":       "Used in Production",
    "damaged":    "Damaged / Scrapped",
    "correction": "Count Correction",
    "returned":   "Returned to Supplier",
    "other":      "Other",
}


def get_category_markups(db: Session) -> dict:
    """Return {category_value: markup_percent} from app_settings."""
    rows = db.query(AppSetting).filter(AppSetting.key.like("markup.%")).all()
    result = {cat.value: 0.0 for cat in InventoryCategory}
    for row in rows:
        cat = row.key.replace("markup.", "")
        try:
            result[cat] = float(row.value or 0)
        except ValueError:
            pass
    return result


def next_sku(db: Session) -> str:
    last = db.query(InventoryItem.sku).order_by(InventoryItem.sku.desc()).first()
    if last and last[0]:
        try:
            num = int(last[0].split("-")[-1]) + 1
        except Exception:
            num = 1
    else:
        num = 1
    return f"VBS-INV-{num:05d}"


# ── Search (HTMX autocomplete for order line items) ───────────────────────
@router.get("/search", response_class=HTMLResponse)
async def search_inventory(
    q: str = "",
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user or len(q.strip()) < 1:
        return HTMLResponse("")
    items = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        (InventoryItem.name.ilike(f"%{q}%") | InventoryItem.sku.ilike(f"%{q}%")),
    ).order_by(InventoryItem.name).limit(10).all()

    if not items:
        return HTMLResponse('<div class="px-3 py-2 text-sm text-gray-400">No items found</div>')

    markups = get_category_markups(db)

    rows = ""
    for item in items:
        cost = item.cost_per_unit or 0
        markup_pct = markups.get(item.category.value, 0)
        sell_price = round(cost * (1 + markup_pct / 100), 2) if cost else ""
        price_str = f"${cost:.2f}/{item.unit}" if cost else item.unit
        qty = item.quantity_on_hand or 0
        qty_str = f"{int(qty) if qty == int(qty) else qty} {item.unit} on hand"
        low = item.reorder_threshold is not None and qty <= item.reorder_threshold
        qty_color = "text-red-500" if low else "text-gray-400"
        rows += (
            f'<div class="px-3 py-2 hover:bg-steel-light cursor-pointer text-sm flex justify-between items-center"'
            f' data-inv-id="{item.id}"'
            f' data-inv-name="{item.name}"'
            f' data-inv-sku="{item.sku or ""}"'
            f' data-inv-unit="{item.unit}"'
            f' data-inv-cost="{item.cost_per_unit or ""}"'
            f' data-inv-price="{sell_price}"'
            f' data-inv-category="{item.category.value}">'
            f'  <span><span class="font-mono text-xs text-gray-400 mr-2">{item.sku or ""}</span>{item.name}</span>'
            f'  <span class="text-xs text-right"><span class="{qty_color} block">{qty_str}</span><span class="text-gray-400">{price_str} cost</span></span>'
            f'</div>'
        )
    return HTMLResponse(rows)


# ── List ──────────────────────────────────────────────────────────────────
@router.get("", response_class=HTMLResponse)
async def list_inventory(
    request: Request,
    category: str = "",
    stock_status: str = "",   # "" | "ok" | "low" | "out"
    q: str = "",
    sort_by: str = "",
    sort_dir: str = "asc",
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)

    query = db.query(InventoryItem).filter(InventoryItem.is_active == True)

    if category:
        query = query.filter(InventoryItem.category == category)
    if q:
        query = query.filter(
            InventoryItem.name.ilike(f"%{q}%") | InventoryItem.sku.ilike(f"%{q}%")
        )

    sort_cols = {
        "sku":      InventoryItem.sku,
        "name":     InventoryItem.name,
        "category": InventoryItem.category,
        "location": InventoryItem.location,
        "on_hand":  InventoryItem.quantity_on_hand,
        "reorder":  InventoryItem.reorder_threshold,
        "cost":     InventoryItem.cost_per_unit,
    }
    if sort_by in sort_cols:
        col = sort_cols[sort_by]
        query = query.order_by(col.desc() if sort_dir == "desc" else col.asc())
    else:
        query = query.order_by(InventoryItem.category, InventoryItem.name)

    items = query.all()

    # Stock status filter (Python-side — uses computed threshold comparison)
    if stock_status == "low":
        items = [i for i in items if i.reorder_threshold is not None and i.quantity_on_hand <= i.reorder_threshold]
    elif stock_status == "out":
        items = [i for i in items if (i.quantity_on_hand or 0) <= 0]
    elif stock_status == "ok":
        items = [i for i in items if (i.quantity_on_hand or 0) > 0 and not (i.reorder_threshold is not None and i.quantity_on_hand <= i.reorder_threshold)]

    low_stock_count = sum(
        1 for i in db.query(InventoryItem).filter(InventoryItem.is_active == True).all()
        if i.reorder_threshold is not None and i.quantity_on_hand <= i.reorder_threshold
    )

    ctx = {
        "request":          request,
        "user":             user,
        "can_see_financials": financials_visible(user),
        "items":            items,
        "category_filter":  category,
        "stock_status":     stock_status,
        "search":           q,
        "sort_by":          sort_by,
        "sort_dir":         sort_dir,
        "low_stock_count":  low_stock_count,
        "category_labels":  CATEGORY_LABELS,
        "categories":       list(InventoryCategory),
    }

    # Return only table rows for HTMX requests
    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("inventory/_rows.html", ctx)

    return templates.TemplateResponse("inventory/list.html", ctx)


# ── New ───────────────────────────────────────────────────────────────────
@router.get("/new", response_class=HTMLResponse)
async def new_item_form(
    request: Request,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    return templates.TemplateResponse("inventory/new.html", {
        "request":         request,
        "user":            user,
        "can_see_financials": financials_visible(user),
        "categories":      list(InventoryCategory),
        "category_labels": CATEGORY_LABELS,
        "next_sku":        next_sku(db),
    })


@router.post("/new", response_class=HTMLResponse)
async def create_item(
    request: Request,
    sku: str                    = Form(...),
    name: str                   = Form(...),
    category: str               = Form(...),
    description: Optional[str]  = Form(None),
    unit: str                   = Form(...),
    quantity_on_hand: float     = Form(0),
    reorder_threshold: Optional[float] = Form(None),
    cost_per_unit: Optional[float]     = Form(None),
    location: Optional[str]     = Form(None),
    supplier_name: Optional[str] = Form(None),
    supplier_contact: Optional[str] = Form(None),
    notes: Optional[str]        = Form(None),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)

    item = InventoryItem(
        sku=sku,
        name=name,
        category=InventoryCategory(category),
        description=description,
        unit=unit,
        quantity_on_hand=quantity_on_hand,
        reorder_threshold=reorder_threshold,
        cost_per_unit=cost_per_unit,
        location=location,
        supplier_name=supplier_name,
        supplier_contact=supplier_contact,
        notes=notes,
    )
    db.add(item)
    db.flush()

    # Log initial stock if > 0
    if quantity_on_hand and quantity_on_hand > 0:
        adj = InventoryAdjustment(
            item_id=item.id,
            delta=quantity_on_hand,
            reason=AdjustmentReason.received,
            notes="Initial stock on hand",
            recorded_by_id=user.id,
        )
        db.add(adj)

    db.commit()
    return RedirectResponse(f"/inventory/{item.id}", status_code=302)


# ── Markup Settings ───────────────────────────────────────────────────────
@router.get("/settings", response_class=HTMLResponse)
async def markup_settings_page(request: Request, user=Depends(get_current_user), db: Session = Depends(get_db)):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    markups = get_category_markups(db)
    return templates.TemplateResponse("inventory/settings.html", {
        "request": request, "user": user,
        "can_see_financials": financials_visible(user),
        "markups": markups,
        "category_labels": CATEGORY_LABELS,
        "categories": list(InventoryCategory),
        "saved": request.query_params.get("saved") == "1",
    })


@router.post("/settings")
async def save_markup_settings(
    request: Request,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    form = await request.form()
    for cat in InventoryCategory:
        key = f"markup.{cat.value}"
        val = str(form.get(f"markup_{cat.value}", "0")).strip()
        try:
            float(val)
        except ValueError:
            val = "0"
        setting = db.query(AppSetting).filter(AppSetting.key == key).first()
        if setting:
            setting.value = val
        else:
            db.add(AppSetting(key=key, value=val))
    db.commit()
    return RedirectResponse("/inventory/settings?saved=1", status_code=302)


# ── Excel helpers ─────────────────────────────────────────────────────────
IMPORT_COLS = [
    "SKU", "Name", "Category", "Unit",
    "Quantity On Hand", "Reorder At", "Unit Cost ($)", "Weight Per Unit (lbs)",
    "Retail Markup",
    "Location", "Supplier Name", "Supplier Contact", "Notes", "Description",
]
AUDIT_COLS = [
    "SKU", "Name", "Category", "Location",
    "System Qty", "Counted Qty", "Unit", "Notes",
]
HDR_FILL  = PatternFill("solid", fgColor="1B3A4B")
HDR_FONT  = Font(bold=True, color="FFFFFF")
LOCK_FILL = PatternFill("solid", fgColor="F0F4F7")
LOCK_FONT = Font(color="888888")

def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 20


# ── Template downloads ─────────────────────────────────────────────────────
@router.get("/template/import")
async def download_import_template(user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Import data ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Inventory Import"
    for c, h in enumerate(IMPORT_COLS, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, len(IMPORT_COLS))
    # Example row
    example = [
        "", '1/2" A36 HR Steel Plate', "steel_plate_a36", "ea",
        500, 100, 281.89, 489.92, 3.0, "Plate / Rack A", "Steel Supply Co.", "555-1234", "", "",
    ]
    for c, v in enumerate(example, 1):
        ws.cell(row=2, column=c, value=v)
    # Helper note
    ws.cell(row=3, column=1, value="← Delete example row before importing. Leave SKU blank to auto-generate. See 'Categories' tab for valid category codes.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="888888")
    ws.merge_cells(f"A3:{get_column_letter(len(IMPORT_COLS))}3")
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25

    # ── Sheet 2: Category reference ────────────────────────────────────────
    ws2 = wb.create_sheet("Categories")
    ref_hdr_fill = PatternFill("solid", fgColor="1B3A4B")
    ref_hdr_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="E8F0F5")
    section_font = Font(bold=True, color="1B3A4B")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 45

    for col, hdr in enumerate(["Category Code", "Display Name", "What goes here"], 1):
        cell = ws2.cell(row=1, column=col, value=hdr)
        cell.fill = ref_hdr_fill
        cell.font = ref_hdr_font
        cell.alignment = Alignment(horizontal="center")

    rows = [
        # (section_heading, code, label, description)
        ("Carbon Steel — Tube & Pipe", None, None, None),
        (None, "steel_pipe",           "Steel — Pipe (Sched 40)",          "Schedule 40 Pipe"),
        (None, "steel_rect_tube",      "Steel — Rect Tube",                "Rectangular hollow structural tubing"),
        (None, "steel_sq_tube",        "Steel — Sq Tube",                  "Square hollow structural tubing"),
        (None, "steel_dom_tube",       "Steel — DOM Tube",                 "DOM (drawn over mandrel) round tube"),
        ("Carbon Steel — Structural",  None, None, None),
        (None, "steel_channel",        "Steel — Channel",                  "C-channel / structural channel"),
        (None, "steel_angle",          "Steel — Angle (Equal)",            "Equal leg angle iron"),
        (None, "steel_angle_unequal",  "Steel — Angle (Unequal)",          "Unequal leg angle iron"),
        (None, "steel_ibeam",          "Steel — I-Beam",                   "Standard I-beams (S-series)"),
        (None, "steel_wide_flange",    "Steel — Wide Flange",              "W-series wide flange beams"),
        (None, "steel_tstock",         "Steel — T-Stock",                  "T-shaped structural stock"),
        (None, "steel_columns",        "Steel — Columns",                  "Columns (N/S)"),
        ("Carbon Steel — Bar",         None, None, None),
        (None, "steel_flat_bar",       "Steel — Flat Bar",                 "HR/CR flat bar"),
        (None, "steel_round_bar",      "Steel — Round Bar",                "HR/CR round bar"),
        (None, "steel_square_bar",     "Steel — Square Bar",               "HR/CR square bar"),
        (None, "steel_strip_hr",       "Steel — HR Strip",                 "Hot-rolled dry strip"),
        ("Carbon Steel — Plate & Sheet", None, None, None),
        (None, "steel_plate_a36",      "Steel — Plate A36",                "A36 structural plate"),
        (None, "steel_plate_ar400",    "Steel — Plate AR400",              "AR400 abrasion-resistant plate"),
        (None, "steel_floor_plate",    "Steel — Floor Plate",              "Diamond / tread plate"),
        (None, "steel_sheet_hr",       "Steel — Sheet HR",                 "Hot-rolled sheet"),
        (None, "steel_sheet_cr",       "Steel — Sheet CR",                 "Cold-rolled sheet"),
        (None, "steel_sheet_galv",     "Steel — Sheet Galvanized",         "Galvanized sheet"),
        (None, "steel_sheet_perf",     "Steel — Sheet Perforated",         "Perforated sheet"),
        ("Carbon Steel — Misc",        None, None, None),
        (None, "steel_expanded",       "Steel — Expanded Metal",           "Expanded metal mesh"),
        (None, "steel_grip_strut",     "Steel — Grip Strut",               "Steel grip strut / safety grating"),
        (None, "steel_bar_grating",    "Steel — Bar Grating",              "Welded bar grating panels"),
        (None, "steel_decking",        "Steel — Floor Decking",            "Corrugated floor decking"),
        (None, "steel_rebar",          "Steel — Rebar",                    "Deformed reinforcing bar"),
        ("Aluminum — Structural",      None, None, None),
        (None, "alum_angle",           "Alum — Angle (Equal)",             "Aluminum equal leg angle"),
        (None, "alum_angle_unequal",   "Alum — Angle (Unequal)",           "Aluminum unequal leg angle"),
        (None, "alum_channel",         "Alum — Channel",                   "Aluminum channel"),
        (None, "alum_flat_bar",        "Alum — Flat Bar",                  "Aluminum flat bar"),
        (None, "alum_round",           "Alum — Round Bar",                 "Aluminum round bar"),
        (None, "alum_square_bar",      "Alum — Square Bar",                "Aluminum square bar"),
        (None, "alum_sq_tube",         "Alum — Sq Tube",                   "Aluminum square tube"),
        (None, "alum_grip_strut",      "Alum — Grip Strut",                "Aluminum grip strut / safety grating"),
        ("Aluminum — Sheet",           None, None, None),
        (None, "alum_sheet",           "Alum — Sheet",                     "Aluminum sheet (3003 / 5052 / 6061)"),
        (None, "alum_treadbrite",      "Alum — Tread Brite",               "Aluminum tread brite sheet"),
        ("Stainless",                  None, None, None),
        (None, "ss_round_bar",         "Stainless — Round Bar",            "Stainless round bar"),
        (None, "ss_square_bar",        "Stainless — Square Bar",           "Stainless square bar"),
        (None, "ss_sheet",             "Stainless — Sheet",                "SS Sheet — #4 / 2B / STD"),
        ("Hardware",                   None, None, None),
        (None, "hardware_fasteners",   "Hardware — Fasteners",             "Nuts, washers, anchors, threaded rod, studs, pipe elbows"),
        (None, "hardware_caps",        "Hardware — Caps",                  "Square steel caps, dome caps"),
        (None, "hardware_gussets",     "Hardware — Gussets",               "Gusset plates"),
        (None, "hardware_base_plates", "Hardware — Base Plates",           "Base plates"),
        (None, "hardware_handrail",    "Hardware — Handrail",              "Handrail covers, elbows (90°/55°/35°)"),
        (None, "hardware_hinges",      "Hardware — Hinges",                "Butt hinges, piano hinges"),
        ("Other",                      None, None, None),
        (None, "bumper_posts",         "Bumper Posts",                     "Bumper Posts — L-D / H-D"),
        (None, "consumables",          "Consumables",                      "Welding wire, gas, grinding discs, bandsaw blades, paint drums"),
        (None, "retail",               "Retail / Walk-In",                 "Walk-in retail items (3× cost markup applied automatically)"),
    ]

    r = 2
    for (section, code, label, desc) in rows:
        if section:
            # Section header row
            cell = ws2.cell(row=r, column=1, value=section)
            cell.fill = section_fill
            cell.font = section_font
            ws2.merge_cells(f"A{r}:C{r}")
        else:
            ws2.cell(row=r, column=1, value=code).font = Font(name="Courier New", size=10)
            ws2.cell(row=r, column=2, value=label)
            ws2.cell(row=r, column=3, value=desc).font = Font(color="555555", italic=True)
        r += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vbs_inventory_import_template.xlsx"})


@router.get("/template/audit")
async def download_audit_template(user=Depends(get_current_user), db: Session = Depends(get_db)):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    items = (db.query(InventoryItem)
               .filter(InventoryItem.is_active == True)
               .order_by(InventoryItem.category, InventoryItem.name)
               .all())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Physical Count"
    for c, h in enumerate(AUDIT_COLS, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, len(AUDIT_COLS))
    for r, item in enumerate(items, 2):
        vals = [
            item.sku, item.name,
            CATEGORY_LABELS.get(item.category.value, item.category.value),
            item.location or "",
            item.quantity_on_hand, None,   # Counted Qty blank for user
            item.unit, "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c != 6:   # lock everything except Counted Qty
                cell.fill = LOCK_FILL
                cell.font = LOCK_FONT
    # Bold the Counted Qty header to call it out
    ws.cell(row=1, column=6).font = Font(bold=True, color="FFAA00")
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"vbs_physical_count_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Import page ────────────────────────────────────────────────────────────
@router.get("/import", response_class=HTMLResponse)
async def import_page(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    return templates.TemplateResponse("inventory/import.html", {
        "request": request, "user": user,
        "can_see_financials": financials_visible(user),
    })


@router.post("/import", response_class=HTMLResponse)
async def process_import(
    request: Request,
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception:
        return templates.TemplateResponse("inventory/import.html", {
            "request": request, "user": user,
            "can_see_financials": financials_visible(user),
            "error": "Could not read file — make sure it's a valid .xlsx file.",
        })

    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    created, updated, errors = [], [], []
    CATS = {c.value for c in InventoryCategory}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rd = dict(zip(headers, row))

        name = str(rd.get("Name") or "").strip()
        if not name or name.lower().startswith("←") or name.lower().startswith("delete"):
            continue

        cat_raw = str(rd.get("Category") or "").strip().lower()
        if cat_raw not in CATS:
            errors.append(f"Row {row_idx} ({name or '—'}): invalid category '{cat_raw}'. "
                          f"Must be: {', '.join(sorted(CATS))}")
            continue

        unit = str(rd.get("Unit") or "each").strip() or "each"

        def _float(key):
            v = rd.get(key)
            try:
                return float(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None

        qty       = _float("Quantity On Hand") or 0.0
        reorder   = _float("Reorder At")
        cost      = _float("Unit Cost ($)")
        weight    = _float("Weight Per Unit (lbs)")
        markup    = _float("Retail Markup") or 3.0
        location  = str(rd.get("Location") or "").strip() or None
        sup_name  = str(rd.get("Supplier Name") or "").strip() or None
        sup_con   = str(rd.get("Supplier Contact") or "").strip() or None
        notes_val = str(rd.get("Notes") or "").strip() or None
        desc_val  = str(rd.get("Description") or "").strip() or None
        sku       = str(rd.get("SKU") or "").strip()

        existing = db.query(InventoryItem).filter(InventoryItem.sku == sku).first() if sku else None

        if existing:
            existing.name              = name
            existing.category          = InventoryCategory(cat_raw)
            existing.unit              = unit
            existing.quantity_on_hand  = qty
            existing.reorder_threshold = reorder
            existing.cost_per_unit     = cost
            existing.weight_per_unit   = weight
            existing.retail_markup     = markup
            existing.location          = location
            existing.supplier_name     = sup_name
            existing.supplier_contact  = sup_con
            existing.notes             = notes_val
            existing.description       = desc_val
            updated.append({"sku": sku, "name": name})
        else:
            if not sku:
                sku = next_sku(db)
            item = InventoryItem(
                sku=sku, name=name,
                category=InventoryCategory(cat_raw),
                unit=unit, quantity_on_hand=qty,
                reorder_threshold=reorder, cost_per_unit=cost,
                weight_per_unit=weight, retail_markup=markup,
                location=location, supplier_name=sup_name,
                supplier_contact=sup_con, notes=notes_val,
                description=desc_val,
            )
            db.add(item)
            db.flush()
            if qty > 0:
                db.add(InventoryAdjustment(
                    item_id=item.id, delta=qty,
                    reason=AdjustmentReason.received,
                    notes="Initial stock — bulk import",
                    recorded_by_id=user.id,
                ))
            created.append({"sku": sku, "name": name})

    db.commit()
    return templates.TemplateResponse("inventory/import_result.html", {
        "request": request, "user": user,
        "can_see_financials": financials_visible(user),
        "created": created, "updated": updated, "errors": errors,
    })


# ── Audit / Physical Count ─────────────────────────────────────────────────
@router.get("/audit", response_class=HTMLResponse)
async def audit_page(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)
    return templates.TemplateResponse("inventory/audit.html", {
        "request": request, "user": user,
        "can_see_financials": financials_visible(user),
    })


@router.post("/audit", response_class=HTMLResponse)
async def process_audit(
    request: Request,
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not user:
        return RedirectResponse("/auth/login", status_code=302)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception:
        return templates.TemplateResponse("inventory/audit.html", {
            "request": request, "user": user,
            "can_see_financials": financials_visible(user),
            "error": "Could not read file — make sure it's a valid .xlsx file.",
        })

    ws = wb.active
    # Counted Qty is always column 6 (index 5) in our template
    adjusted, no_change, not_found, errors = [], [], [], []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sku = str(row[0] or "").strip() if row[0] is not None else ""
        if not sku:
            continue
        counted_raw = row[5]   # column F = Counted Qty
        if counted_raw is None or counted_raw == "":
            continue
        try:
            counted = float(counted_raw)
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx} ({sku}): '{counted_raw}' is not a valid number")
            continue

        item = db.query(InventoryItem).filter(InventoryItem.sku == sku).first()
        if not item:
            not_found.append(sku)
            continue

        old_qty = item.quantity_on_hand or 0.0
        if abs(counted - old_qty) < 0.0001:
            no_change.append({"sku": sku, "name": item.name, "qty": old_qty})
            continue

        delta = counted - old_qty
        notes_raw = str(row[7] or "").strip() if len(row) > 7 else ""
        audit_note = f"Physical count — system: {old_qty} {item.unit}, counted: {counted} {item.unit}."
        if notes_raw:
            audit_note += f" Note: {notes_raw}"

        item.quantity_on_hand = 